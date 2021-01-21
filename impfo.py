import shutil
from urllib.request import Request, urlopen, urlretrieve
from io import BytesIO
from openpyxl import load_workbook
import  datetime

XLSX_URL = "https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Impfquotenmonitoring.xlsx?__blob=publicationFile"
PATH = "data.xlsx"

request = Request(XLSX_URL, headers={'User-Agent': 'Mozilla/5.0'})
f = urlopen(request).read()

wb = load_workbook(filename=BytesIO(f), read_only=True, data_only=True)

vaccination_sheet = wb["Impfungen_proTag"]

for row in vaccination_sheet.iter_rows(2):
    if isinstance(row[0].value, datetime.date):
        print(row[0].value.strftime("%Y-%m-%d") + "\t" + str(row[1].value) + "\t" + str(row[2].value))
    elif row[0].value != None:
        print(str(row[0].value) + "\t" + str(row[1].value) + "\t" + str(row[2].value))
wb.close()
